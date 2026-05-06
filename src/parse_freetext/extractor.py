from __future__ import annotations

from collections.abc import Iterable, Sequence
import csv
from dataclasses import dataclass
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet


class ExtractionError(ValueError):
    """Raised when spreadsheet extraction cannot continue safely."""


@dataclass(frozen=True)
class ExtractionResult:
    rows_seen: int
    files_written: int
    files_skipped: int
    output_dir: Path


@dataclass(frozen=True)
class SheetInfo:
    name: str
    headers: list[str]


def inspect_spreadsheet(input_file: Path) -> list[SheetInfo]:
    """Return sheet names and first-row headers for a spreadsheet-like input."""

    if not input_file.exists():
        raise ExtractionError(f"Input file does not exist: {input_file}")
    if not input_file.is_file():
        raise ExtractionError(f"Input path is not a file: {input_file}")
    suffix = _supported_suffix(input_file)

    if suffix in {".csv", ".tsv"}:
        rows = _read_delimited_rows(input_file, suffix)
        header_row = rows[0] if rows else []
        return [SheetInfo(name=input_file.stem, headers=_headers_from_row(header_row))]

    workbook = load_workbook(input_file, read_only=True, data_only=True)
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            header_row = next(rows, ())
            headers = _headers_from_row(header_row)
            sheets.append(SheetInfo(name=worksheet.title, headers=headers))
        return sheets
    finally:
        workbook.close()


def extract_text_files(
    input_file: Path,
    output_dir: Path,
    sheets: Sequence[str] | None,
    transaction_id_column: str,
    text_columns: Sequence[str],
    *,
    overwrite: bool = False,
    append_sheet_name: bool = False,
) -> ExtractionResult:
    """Extract selected spreadsheet columns into record-named text files."""

    if not input_file.exists():
        raise ExtractionError(f"Input file does not exist: {input_file}")
    if not input_file.is_file():
        raise ExtractionError(f"Input path is not a file: {input_file}")
    if not text_columns:
        raise ExtractionError("At least one --text-column value is required.")
    suffix = _supported_suffix(input_file)

    if suffix in {".csv", ".tsv"}:
        return _extract_delimited_text_files(
            input_file=input_file,
            output_dir=output_dir,
            sheets=sheets,
            transaction_id_column=transaction_id_column,
            text_columns=text_columns,
            overwrite=overwrite,
        )

    return _extract_xlsx_text_files(
        input_file=input_file,
        output_dir=output_dir,
        sheets=sheets,
        transaction_id_column=transaction_id_column,
        text_columns=text_columns,
        overwrite=overwrite,
        append_sheet_name=append_sheet_name,
    )


def _extract_xlsx_text_files(
    input_file: Path,
    output_dir: Path,
    sheets: Sequence[str] | None,
    transaction_id_column: str,
    text_columns: Sequence[str],
    *,
    overwrite: bool,
    append_sheet_name: bool,
) -> ExtractionResult:
    workbook = load_workbook(input_file, read_only=True, data_only=True)
    try:
        selected_sheet_names = _resolve_sheet_names(workbook.sheetnames, sheets)
        output_dir.mkdir(parents=True, exist_ok=True)

        rows_seen = 0
        files_written = 0
        files_skipped = 0

        for sheet_name in selected_sheet_names:
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                continue

            header = _normalize_header(header_row)
            transaction_idx = _resolve_column(transaction_id_column, header, worksheet)
            text_indexes = [_resolve_column(column, header, worksheet) for column in text_columns]

            for row_number, row in enumerate(rows, start=2):
                rows_seen += 1
                transaction_id = _cell_to_text(_value_at(row, transaction_idx))
                if not transaction_id:
                    files_skipped += 1
                    continue

                text_parts = [
                    _cell_to_text(_value_at(row, text_idx))
                    for text_idx in text_indexes
                ]
                text = "\n\n".join(part for part in text_parts if part)
                if not text:
                    files_skipped += 1
                    continue

                filename_stem = _safe_filename(transaction_id)
                if append_sheet_name:
                    filename_stem = f"{filename_stem}_{_safe_filename(sheet_name)}"

                output_file = output_dir / f"{filename_stem}.txt"
                if output_file.exists() and not overwrite:
                    raise ExtractionError(
                        f"Output file already exists for row {row_number} on sheet "
                        f"{sheet_name!r}: {output_file}. Use --overwrite to replace it "
                        "or --append-sheet-name to avoid cross-sheet collisions."
                    )

                output_file.write_text(text + "\n", encoding="utf-8")
                files_written += 1

        return ExtractionResult(rows_seen, files_written, files_skipped, output_dir)
    finally:
        workbook.close()


def _extract_delimited_text_files(
    input_file: Path,
    output_dir: Path,
    sheets: Sequence[str] | None,
    transaction_id_column: str,
    text_columns: Sequence[str],
    *,
    overwrite: bool,
) -> ExtractionResult:
    if sheets:
        raise ExtractionError("--sheet is only supported for .xlsx workbooks, not CSV/TSV files.")

    rows = _read_delimited_rows(input_file, input_file.suffix.lower())
    if not rows:
        return ExtractionResult(0, 0, 0, output_dir)

    header_row = rows[0]
    header = _normalize_header(header_row)
    transaction_idx = _resolve_column(transaction_id_column, header, input_file.stem)
    text_indexes = [_resolve_column(column, header, input_file.stem) for column in text_columns]

    output_dir.mkdir(parents=True, exist_ok=True)
    rows_seen = 0
    files_written = 0
    files_skipped = 0

    for row_number, row in enumerate(rows[1:], start=2):
        rows_seen += 1
        transaction_id = _cell_to_text(_value_at(row, transaction_idx))
        if not transaction_id:
            files_skipped += 1
            continue

        text_parts = [
            _cell_to_text(_value_at(row, text_idx))
            for text_idx in text_indexes
        ]
        text = "\n\n".join(part for part in text_parts if part)
        if not text:
            files_skipped += 1
            continue

        output_file = output_dir / f"{_safe_filename(transaction_id)}.txt"
        if output_file.exists() and not overwrite:
            raise ExtractionError(
                f"Output file already exists for row {row_number}: {output_file}. "
                "Use --overwrite to replace it."
            )

        output_file.write_text(text + "\n", encoding="utf-8")
        files_written += 1

    return ExtractionResult(rows_seen, files_written, files_skipped, output_dir)


def _supported_suffix(input_file: Path) -> str:
    suffix = input_file.suffix.lower()
    if suffix not in {".xlsx", ".csv", ".tsv"}:
        raise ExtractionError("input_file must be a supported spreadsheet file: .xlsx, .csv, or .tsv")
    return suffix


def _read_delimited_rows(input_file: Path, suffix: str) -> list[list[str]]:
    delimiter = "\t" if suffix == ".tsv" else ","
    with input_file.open(encoding="utf-8-sig", newline="") as file:
        return [row for row in csv.reader(file, delimiter=delimiter)]


def _resolve_sheet_names(available: Sequence[str], requested: Sequence[str] | None) -> list[str]:
    if not requested:
        return list(available)

    missing = [sheet for sheet in requested if sheet not in available]
    if missing:
        available_text = ", ".join(available)
        missing_text = ", ".join(missing)
        raise ExtractionError(
            f"Unknown sheet(s): {missing_text}. Available sheets: {available_text}"
        )
    return list(requested)


def _normalize_header(header_row: Iterable[object]) -> dict[str, int]:
    header: dict[str, int] = {}
    for index, value in enumerate(header_row):
        if value is None:
            continue
        key = str(value).strip()
        if key:
            header[key.casefold()] = index
    return header


def _resolve_column(column: str, header: dict[str, int], source: Worksheet | str) -> int:
    value = column.strip()
    if not value:
        raise ExtractionError("Column values cannot be empty.")

    if value.casefold() in header:
        return header[value.casefold()]

    if value.isdigit():
        index = int(value) - 1
        if index < 0:
            raise ExtractionError(f"Column numbers are 1-based: {column!r}")
        return index

    if re.fullmatch(r"[A-Za-z]+", value):
        return column_index_from_string(value.upper()) - 1

    header_text = ", ".join(sorted(header)) or "none"
    source_name = source.title if isinstance(source, Worksheet) else source
    raise ExtractionError(
        f"Column {column!r} was not found on {source_name!r}. "
        f"Known headers: {header_text}"
    )


def _value_at(row: Sequence[object], index: int) -> object | None:
    if index >= len(row):
        return None
    return row[index]


def _cell_to_text(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_filename(value: str) -> str:
    filename = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    filename = filename.strip("._")
    if not filename:
        raise ExtractionError(f"Record id {value!r} cannot be used as a filename.")
    return filename[:200]


def _headers_from_row(header_row: Iterable[object]) -> list[str]:
    return [
        str(value).strip()
        for value in header_row
        if value is not None and str(value).strip()
    ]
